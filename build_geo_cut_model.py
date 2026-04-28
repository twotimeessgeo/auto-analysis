#!/usr/bin/env python3
"""Build the geography cut prediction model from grading workbooks and EBSi data."""

from __future__ import annotations

import argparse
import json
import math
import re
import warnings
from collections import Counter
from datetime import datetime
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


SUBJECTS = ("한국지리", "세계지리")
CUTS = ("1", "2", "3")
CUT_FEATURES = ("mean", "nat_sd", "hard15_rate", "hard15_sd", "under40_points", "easy5_rate")
SD_FEATURES = ("mean", "hard15_rate", "hard15_sd", "under40_points", "under50_points", "easy5_rate")
RIDGE_ALPHA = 1.0
GUESS_RATE = 20.0
LOGIT_EPS = 0.005
DEFAULT_EBSI_MONTHS = ("06", "09", "11")


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def as_float(value: object) -> float | None:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def cut_midpoint(value: object) -> float | None:
    numbers = [float(part) for part in re.findall(r"\d+(?:\.\d+)?", str(value or ""))]
    if not numbers:
        return None
    return sum(numbers) / len(numbers)


def extract_records(source_dir: Path) -> list[dict]:
    records: list[dict] = []
    for workbook_path in sorted(source_dir.glob("*.xlsx")):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))

        for row_index, row in enumerate(rows):
            if len(row) < 4 or row[0] not in SUBJECTS or row[1] != "정답" or row[3] != "정답률":
                continue

            items = []
            for item_row in rows[row_index + 1 : row_index + 21]:
                try:
                    question = int(str(item_row[0]).strip())
                except (TypeError, ValueError):
                    continue
                items.append(
                    {
                        "question": question,
                        "points": as_float(item_row[2]),
                        "academy_rate": as_float(item_row[3]),
                    }
                )

            summary: dict[str, float] = {}
            for summary_row in rows[row_index + 21 : min(row_index + 32, len(rows))]:
                if not summary_row:
                    continue
                for col_index, value in enumerate(summary_row):
                    if not isinstance(value, str):
                        continue
                    label = value.strip()
                    next_value = summary_row[col_index + 1] if col_index + 1 < len(summary_row) else None
                    raw_value = summary_row[col_index + 2] if col_index + 2 < len(summary_row) else None
                    if label == "평균":
                        summary["academy_mean"] = as_float(next_value)
                    elif label == "표준편차":
                        summary["academy_sd"] = as_float(next_value)
                    elif label == "응시자수":
                        summary["sample_size"] = as_float(next_value)
                    elif label == "전국1컷":
                        summary["std1"] = as_float(next_value)
                        summary["raw1"] = cut_midpoint(raw_value)
                    elif label == "전국2컷":
                        summary["std2"] = as_float(next_value)
                        summary["raw2"] = cut_midpoint(raw_value)
                    elif label == "전국3컷":
                        summary["std3"] = as_float(next_value)
                        summary["raw3"] = cut_midpoint(raw_value)

            required = ("raw1", "raw2", "raw3", "std1", "std2", "std3")
            if len(items) != 20 or any(summary.get(key) is None for key in required):
                continue

            raw_scores = np.array([summary["raw1"], summary["raw2"], summary["raw3"]], dtype=float)
            standard_scores = np.array([summary["std1"], summary["std2"], summary["std3"]], dtype=float)
            slope, intercept = np.polyfit(raw_scores, standard_scores, 1)
            summary["national_sd"] = 10 / slope
            summary["national_mean"] = (50 - intercept) / slope
            summary["national_rate"] = summary["national_mean"] / 50 * 100

            records.append(
                {
                    "source_file": workbook_path.name,
                    "subject": row[0],
                    "items": items,
                    **summary,
                }
            )
    return records


def weighted_academy_rate(record: dict) -> float:
    return sum(item["points"] * item["academy_rate"] / 100 for item in record["items"]) / 50 * 100


def sigmoid(value: float) -> float:
    return 1 / (1 + math.exp(-value))


def chance_floor_logit(rate: float, floor: float = GUESS_RATE) -> float:
    scaled = (rate - floor) / (100 - floor)
    scaled = min(1 - LOGIT_EPS, max(LOGIT_EPS, scaled))
    return math.log(scaled / (1 - scaled))


def chance_floor_rate(value: float, floor: float = GUESS_RATE) -> float:
    return floor + (100 - floor) * sigmoid(value)


def fit_academy_to_national(records: list[dict]) -> dict[str, dict]:
    relation = {}
    for subject in SUBJECTS:
        subject_records = [record for record in records if record["subject"] == subject]
        academy_rates = np.array([weighted_academy_rate(record) for record in subject_records], dtype=float)
        national_rates = np.array([record["national_rate"] for record in subject_records], dtype=float)
        slope, intercept = np.polyfit(academy_rates, national_rates, 1)
        predictions = np.array(
            [min(99.5, max(GUESS_RATE, intercept + slope * rate)) for rate in academy_rates],
            dtype=float,
        )
        rmse = float(np.sqrt(np.mean((predictions - national_rates) ** 2)))
        relation[subject] = {
            "method": "linear_with_chance_floor",
            "chance_floor": GUESS_RATE,
            "intercept": float(intercept),
            "slope": float(slope),
            "rmse_pct_points": rmse,
            "records": len(subject_records),
        }
    return relation


def logit(value: float) -> float:
    value = min(0.995, max(0.005, value))
    return math.log(value / (1 - value))


def convert_academy_rate(rate: float, subject: str, relation: dict[str, dict]) -> float:
    rel = relation[subject]
    if rel.get("method") == "chance_floor_logit":
        return chance_floor_rate(rel["intercept"] + rel["slope"] * chance_floor_logit(rate, rel["chance_floor"]), rel["chance_floor"])
    if rel.get("method") == "linear_with_chance_floor":
        return min(99.5, max(rel["chance_floor"], rel["intercept"] + rel["slope"] * rate))
    return min(99.5, max(0.5, rel["intercept"] + rel["slope"] * rate))


def impute_national_rates(record: dict, relation: dict[str, dict]) -> list[float]:
    rel = relation[record["subject"]]
    points = [item["points"] for item in record["items"]]
    base_logits = []
    for item in record["items"]:
        rate = convert_academy_rate(item["academy_rate"], record["subject"], relation)
        base_logits.append(logit(min(99.5, max(0.5, rate)) / 100))

    target_mean = record["national_mean"] / 50
    low, high = -10.0, 10.0
    for _ in range(70):
        mid = (low + high) / 2
        mean = sum(point * sigmoid(item_logit + mid) for point, item_logit in zip(points, base_logits)) / 50
        if mean < target_mean:
            low = mid
        else:
            high = mid

    offset = (low + high) / 2
    return [100 * sigmoid(item_logit + offset) for item_logit in base_logits]


def ranked_item_features(
    rates: list[float],
    points: list[float],
    count: int = 15,
    easy_rate_fallback: float | None = None,
) -> dict[str, float]:
    ranked_items = sorted(zip(rates, points), key=lambda pair: pair[0])
    hard_items = ranked_items[:count]
    easy_items = ranked_items[count:]
    hard_points = sum(point for _, point in hard_items)
    hard_rate = sum(point * rate for rate, point in hard_items) / hard_points
    hard_variance = sum(
        point * (rate - hard_rate) ** 2 for rate, point in hard_items
    ) / hard_points
    easy_points = sum(point for _, point in easy_items)
    if easy_points:
        easy_rate = sum(point * rate for rate, point in easy_items) / easy_points
    elif easy_rate_fallback is not None:
        easy_rate = easy_rate_fallback
    else:
        easy_rate = max(rates)
    return {
        "hard15_rate": hard_rate,
        "hard15_sd": math.sqrt(hard_variance),
        "under40_points": sum(point for rate, point in hard_items if rate < 40),
        "under50_points": sum(point for rate, point in hard_items if rate < 50),
        "easy5_rate": easy_rate,
    }


def infer_easy_missing_rate(easiest_top15_rate: float) -> float:
    return (100.0 + easiest_top15_rate) / 2


def feature_values(rates: list[float], points: list[float]) -> dict[str, float]:
    probabilities = [rate / 100 for rate in rates]
    mean = sum(point * probability for point, probability in zip(points, probabilities))
    independent_sd = math.sqrt(
        sum(point * point * probability * (1 - probability) for point, probability in zip(points, probabilities))
    )
    weighted_rate = mean / sum(points) * 100
    rate_variance = sum(
        point * (rate - weighted_rate) ** 2 for point, rate in zip(points, rates)
    ) / sum(points)
    features = {
        "mean": mean,
        "ind_sd": independent_sd,
        "rate_sd": math.sqrt(rate_variance),
    }
    features.update(ranked_item_features(rates, points))
    return features


def calibrated_ebsi_top15_features(record: dict, rates: list[float], points: list[float]) -> dict[str, float]:
    easiest_top15_rate = max(rates)
    return ranked_item_features(
        rates,
        points,
        easy_rate_fallback=infer_easy_missing_rate(easiest_top15_rate),
    )


def prepare_training_rows(records: list[dict], relation: dict[str, dict]) -> list[dict]:
    rows = []
    for record in records:
        rates = impute_national_rates(record, relation)
        points = [item["points"] for item in record["items"]]
        features = feature_values(rates, points)
        features["mean"] = record["national_mean"]
        features["nat_sd"] = record["national_sd"]
        rows.append(
            {
                "source": "academy_workbook",
                "subject": record["subject"],
                "source_file": record["source_file"],
                "raw1": record["raw1"],
                "raw2": record["raw2"],
                "raw3": record["raw3"],
                "estimated_national_rates": rates,
                **features,
            }
        )
    return rows


def month_allowed(record: dict, allowed_months: set[str] | None) -> bool:
    return allowed_months is None or str(record.get("month")).zfill(2) in allowed_months


def prepare_ebsi_rows(ebsi_path: Path | None, allowed_months: set[str] | None = None) -> list[dict]:
    if not ebsi_path or not ebsi_path.exists():
        return []

    payload = json.loads(ebsi_path.read_text(encoding="utf-8"))
    rows = []
    for record in payload.get("records", []):
        if record.get("subject") not in SUBJECTS:
            continue
        if not month_allowed(record, allowed_months):
            continue
        wrong_top15 = record.get("wrong_top15", [])
        if len(wrong_top15) != 15:
            continue
        items = record.get("items") or []
        if len(items) == 20 and all(
            item.get("national_rate") is not None and item.get("points") is not None
            for item in items
        ):
            points = [float(item["points"]) for item in items]
            rates = [float(item["national_rate"]) for item in items]
            features = feature_values(rates, points)
        else:
            points = [float(item["points"]) for item in wrong_top15]
            rates = [float(item["correct_rate"]) for item in wrong_top15]
            features = calibrated_ebsi_top15_features(record, rates, points)
        features["mean"] = record["national_mean"]
        features["nat_sd"] = record["national_sd"]
        rows.append(
            {
                "source": "ebsi",
                "subject": record["subject"],
                "source_file": f"EBSi {record.get('exam_year')} {record.get('month')}",
                "raw1": record.get("raw1"),
                "raw2": record.get("raw2"),
                "raw3": record.get("raw3"),
                "exact_top15_rates": [
                    float(item["correct_rate"]) for item in wrong_top15
                ],
                **features,
            }
        )
    return rows


def design_matrix(rows: list[dict], features: tuple[str, ...]) -> np.ndarray:
    return np.array(
        [
            [1.0, 1.0 if row["subject"] == "세계지리" else 0.0]
            + [row[feature] for feature in features]
            for row in rows
        ],
        dtype=float,
    )


def fit_model(rows: list[dict], features: tuple[str, ...], alpha: float, target: str) -> np.ndarray:
    rows = [row for row in rows if row.get(target) is not None]
    x = design_matrix(rows, features)
    y = np.array([row[target] for row in rows], dtype=float)
    penalty = np.eye(x.shape[1])
    penalty[0, 0] = 0
    penalty[1, 1] = 0
    return np.linalg.solve(x.T @ x + alpha * penalty, x.T @ y)


def predict_row(row: dict, coefficients: np.ndarray, features: tuple[str, ...]) -> float:
    x = np.array(
        [1.0, 1.0 if row["subject"] == "세계지리" else 0.0]
        + [row[feature] for feature in features],
        dtype=float,
    )
    return float(x @ coefficients)


def cross_validate(rows: list[dict], features: tuple[str, ...], alpha: float, target: str) -> dict:
    rows = [row for row in rows if row.get(target) is not None]
    residuals = []
    by_subject: dict[str, list[float]] = {subject: [] for subject in SUBJECTS}
    for holdout_index, row in enumerate(rows):
        train = [candidate for index, candidate in enumerate(rows) if index != holdout_index]
        coefficients = fit_model(train, features, alpha, target)
        prediction = predict_row(row, coefficients, features)
        residual = prediction - row[target]
        residuals.append(residual)
        by_subject[row["subject"]].append(residual)

    def stats(values: list[float]) -> dict:
        return {
            "rmse": math.sqrt(sum(value * value for value in values) / len(values)),
            "mae": sum(abs(value) for value in values) / len(values),
            "max_abs_error": max(abs(value) for value in values),
        }

    return {
        **stats(residuals),
        "by_subject": {subject: stats(values) for subject, values in by_subject.items() if values},
        "records": len(rows),
        "by_source": {
            source: sum(row.get("source") == source for row in rows)
            for source in sorted({row.get("source") for row in rows})
        },
    }


def default_points(records: list[dict]) -> dict[str, list[int]]:
    defaults = {}
    for subject in SUBJECTS:
        counters = [Counter() for _ in range(20)]
        patterns = []
        for record in records:
            if record["subject"] != subject:
                continue
            pattern = []
            for index, item in enumerate(record["items"]):
                point = int(item["points"])
                pattern.append(point)
                counters[index][point] += 1
            patterns.append(pattern)
        majority = [counter.most_common(1)[0][0] for counter in counters]
        defaults[subject] = min(
            patterns,
            key=lambda pattern: (
                abs(sum(pattern) - 50),
                sum(point != majority[index] for index, point in enumerate(pattern)),
            ),
        )
    return defaults


def build_item_rate_mapping(
    records: list[dict],
    ebsi_path: Path | None,
    allowed_months: set[str] | None = None,
) -> dict[str, dict]:
    ebsi_payload = {}
    if ebsi_path and ebsi_path.exists():
        ebsi_payload = json.loads(ebsi_path.read_text(encoding="utf-8"))

    pooled_academy_rates = sorted(
        float(item["academy_rate"])
        for record in records
        for item in record["items"]
        if item.get("academy_rate") is not None
    )
    pooled_national_rates = sorted(
        float(row["correct_rate"])
        for record in ebsi_payload.get("records", [])
        if month_allowed(record, allowed_months)
        for row in record.get("wrong_top15", [])
        if row.get("correct_rate") is not None
    )

    mapping = {
        "_pooled": {
            "method": "empirical_quantile",
            "academy_source": "all_geography_grading_workbook_item_rates",
            "national_source": "all_geography_ebsi_wrong_answer_top15_exact",
            "coverage": "wrong_top15_only",
            "usable_for_runtime_academy_conversion": False,
            "academy_rates": pooled_academy_rates,
            "national_rates": pooled_national_rates,
            "counts": {
                "academy_rates": len(pooled_academy_rates),
                "national_rates": len(pooled_national_rates),
            },
        }
    }
    for subject in SUBJECTS:
        academy_rates = sorted(
            float(item["academy_rate"])
            for record in records
            if record["subject"] == subject
            for item in record["items"]
            if item.get("academy_rate") is not None
        )
        national_rates = sorted(
            float(row["correct_rate"])
            for record in ebsi_payload.get("records", [])
            if record.get("subject") == subject
            and month_allowed(record, allowed_months)
            for row in record.get("wrong_top15", [])
            if row.get("correct_rate") is not None
        )
        mapping[subject] = {
            "method": "pooled_empirical_quantile_with_subject_shrinkage",
            "subject_weight": 0.25,
            "academy_source": "grading_workbook_item_rates",
            "national_source": "ebsi_wrong_answer_top15_exact",
            "coverage": "wrong_top15_only",
            "usable_for_runtime_academy_conversion": False,
            "academy_rates": academy_rates,
            "national_rates": national_rates,
            "counts": {
                "academy_rates": len(academy_rates),
                "national_rates": len(national_rates),
            },
        }
    return mapping


def build_historical_anchors(ebsi_path: Path | None, allowed_months: set[str] | None = None) -> list[dict]:
    if not ebsi_path or not ebsi_path.exists():
        return []

    payload = json.loads(ebsi_path.read_text(encoding="utf-8"))
    anchors = []
    for record in payload.get("records", []):
        if record.get("subject") not in SUBJECTS or not month_allowed(record, allowed_months):
            continue
        wrong_top15 = record.get("wrong_top15", [])
        if len(wrong_top15) != 15:
            continue
        top15 = sorted(
            (
                float(item["correct_rate"]),
                float(item["points"]),
                int(item["question"]),
            )
            for item in wrong_top15
        )
        items = record.get("items") or []
        if len(items) == 20 and all(
            item.get("national_rate") is not None and item.get("points") is not None
            for item in items
        ):
            item_rates = [float(item["national_rate"]) for item in items]
            item_points = [float(item["points"]) for item in items]
            features = feature_values(item_rates, item_points)
        else:
            top15_points = sum(point for _, point, _ in top15)
            features = ranked_item_features(
                [rate for rate, _, _ in top15],
                [point for _, point, _ in top15],
                easy_rate_fallback=infer_easy_missing_rate(max(rate for rate, _, _ in top15)),
            )
        anchors.append(
            {
                "subject": record["subject"],
                "exam_year": record.get("exam_year"),
                "month": str(record.get("month")).zfill(2),
                "mean": record.get("national_mean"),
                "rate_mean": features.get("mean", record.get("national_mean")),
                "nat_sd": record.get("national_sd"),
                "hard15_rate": features["hard15_rate"],
                "easy5_rate": features["easy5_rate"],
                "easiest_top15_rate": max(rate for rate, _, _ in top15),
                "under40_points": features["under40_points"],
                "under50_points": features["under50_points"],
                "raw1": record.get("raw1"),
                "raw2": record.get("raw2"),
                "raw3": record.get("raw3"),
            }
        )
    return anchors


def historical_cut_caps_from_anchors(
    subject: str,
    features: dict[str, float],
    anchors: list[dict],
) -> dict[str, float]:
    caps: dict[str, float] = {}
    latest_november_anchor: dict | None = None
    for anchor in anchors:
        if anchor.get("subject") != subject:
            continue
        if (
            anchor.get("mean") is None
            or anchor.get("nat_sd") is None
            or anchor.get("hard15_rate") is None
        ):
            continue
        anchor_mean = float(anchor["mean"])
        anchor_rate_mean = float(anchor.get("rate_mean", anchor_mean))
        anchor_sd = max(0.1, float(anchor["nat_sd"]))
        anchor_easy5_rate = float(anchor.get("easy5_rate", anchor["easiest_top15_rate"]))
        no_easier_than_anchor = (
            features["mean"] <= anchor_rate_mean + 0.5
            and features["hard15_rate"] <= float(anchor["hard15_rate"]) + 1.0
            and features["easy5_rate"] <= anchor_easy5_rate + 1.0
        )
        if not no_easier_than_anchor:
            continue

        if str(anchor.get("month")).zfill(2) == "11" and (
            latest_november_anchor is None
            or int(anchor.get("exam_year") or 0) > int(latest_november_anchor.get("exam_year") or 0)
        ):
            latest_november_anchor = anchor

        for cut in CUTS:
            raw_cut = anchor.get(f"raw{cut}")
            if raw_cut is None:
                continue
            raw_cut = float(raw_cut)
            caps[cut] = min(caps.get(cut, raw_cut), raw_cut)

    if latest_november_anchor:
        anchor_mean = float(latest_november_anchor["mean"])
        anchor_sd = max(0.1, float(latest_november_anchor["nat_sd"]))
        for cut in CUTS:
            raw_cut = latest_november_anchor.get(f"raw{cut}")
            if raw_cut is None:
                continue
            raw_cut = float(raw_cut)
            anchor_z = (raw_cut - anchor_mean) / anchor_sd
            scaled_cap = raw_cut + (features["mean"] - anchor_mean)
            scaled_cap += anchor_z * (features["nat_sd"] - anchor_sd)
            caps[cut] = min(caps.get(cut, raw_cut), raw_cut, scaled_cap)

    return caps


def build_runtime_corrections(
    records: list[dict],
    relation: dict[str, dict],
    sd_coefficients: np.ndarray,
    cut_coefficients: dict[str, np.ndarray],
    anchors: list[dict],
) -> dict[str, dict]:
    residuals: dict[str, dict[str, list[float]]] = {
        subject: {cut: [] for cut in CUTS} for subject in SUBJECTS
    }

    for record in records:
        points = [item["points"] for item in record["items"]]
        rates = [
            convert_academy_rate(item["academy_rate"], record["subject"], relation)
            for item in record["items"]
        ]
        features = feature_values(rates, points)
        sd_row = {"subject": record["subject"], **features}
        features["nat_sd"] = clamp(
            predict_row(sd_row, sd_coefficients, SD_FEATURES),
            0.1,
            25.0,
        )
        cut_caps = historical_cut_caps_from_anchors(record["subject"], features, anchors)
        cut_row = {"subject": record["subject"], **features}
        total_points = sum(points)

        for cut in CUTS:
            target = f"raw{cut}"
            prediction = clamp(
                predict_row(cut_row, cut_coefficients[cut], CUT_FEATURES),
                0,
                total_points,
            )
            historical_cap = cut_caps.get(cut)
            if historical_cap is not None:
                prediction = min(prediction, historical_cap)
            residuals[record["subject"]][cut].append(prediction - record[target])

    by_subject = {}
    by_cut = {}
    for cut in CUTS:
        cut_residuals = [
            residual
            for subject in SUBJECTS
            for residual in residuals[subject][cut]
        ]
        by_cut[cut] = max(0.0, float(sum(cut_residuals) / len(cut_residuals)))

    for subject in SUBJECTS:
        by_subject[subject] = {}
        for cut in CUTS:
            subject_residuals = residuals[subject][cut]
            by_subject[subject][cut] = max(
                0.0,
                float(sum(subject_residuals) / len(subject_residuals)),
            )

    return {
        "academy": {
            "method": "subject_cut_bias_subtraction",
            "source": "grading_workbook_runtime_backtest",
            "applied_after_historical_caps": True,
            "by_cut": by_cut,
            "by_subject": by_subject,
            "records": {
                subject: len(residuals[subject]["1"]) for subject in SUBJECTS
            },
        }
    }


def build_model(
    source_dir: Path,
    ebsi_path: Path | None = None,
    ebsi_months: set[str] | None = set(DEFAULT_EBSI_MONTHS),
) -> dict:
    records = extract_records(source_dir)
    relation = fit_academy_to_national(records)
    academy_rows = prepare_training_rows(records, relation)
    ebsi_rows = prepare_ebsi_rows(ebsi_path, ebsi_months)
    rows = academy_rows + ebsi_rows
    cut_coefficient_names = ("intercept", "world_geography_offset", *CUT_FEATURES)
    sd_coefficient_names = ("intercept", "world_geography_offset", *SD_FEATURES)
    sd_coefficients = fit_model(rows, SD_FEATURES, RIDGE_ALPHA, "nat_sd")
    cut_models = {}
    cut_coefficients_by_cut = {}
    for cut in CUTS:
        target = f"raw{cut}"
        coefficients = fit_model(rows, CUT_FEATURES, RIDGE_ALPHA, target)
        cut_coefficients_by_cut[cut] = coefficients
        cut_models[cut] = {
            "target": target,
            "records": sum(row.get(target) is not None for row in rows),
            "coefficients": {
                name: float(value) for name, value in zip(cut_coefficient_names, coefficients)
            },
            "cross_validation": cross_validate(rows, CUT_FEATURES, RIDGE_ALPHA, target),
        }

    historical_anchors = build_historical_anchors(ebsi_path, ebsi_months)
    runtime_corrections = build_runtime_corrections(
        records,
        relation,
        sd_coefficients,
        cut_coefficients_by_cut,
        historical_anchors,
    )

    return {
        "version": "geo-cut-v8-easy-midpoint-imputation",
        "built_at": datetime.now().isoformat(timespec="seconds"),
        "source_dir": str(source_dir),
        "ebsi_path": str(ebsi_path) if ebsi_path and ebsi_path.exists() else None,
        "ebsi_months": sorted(ebsi_months) if ebsi_months is not None else "all",
        "subjects": list(SUBJECTS),
        "cuts": list(CUTS),
        "features": list(CUT_FEATURES),
        "sd_features": list(SD_FEATURES),
        "ridge_alpha": RIDGE_ALPHA,
        "sd_model": {
            "target": "nat_sd",
            "coefficients": {
                name: float(value) for name, value in zip(sd_coefficient_names, sd_coefficients)
            },
            "cross_validation": cross_validate(rows, SD_FEATURES, RIDGE_ALPHA, "nat_sd"),
        },
        "cut_models": cut_models,
        "academy_to_national_rate": relation,
        "item_rate_mapping": build_item_rate_mapping(records, ebsi_path, ebsi_months),
        "historical_anchors": historical_anchors,
        "runtime_corrections": runtime_corrections,
        "default_points": default_points(records),
        "training_records": {
            "total": len(rows),
            "academy_workbook": len(academy_rows),
            "ebsi": len(ebsi_rows),
            "ebsi_months": sorted(ebsi_months) if ebsi_months is not None else "all",
            "by_subject": {
                subject: sum(row["subject"] == subject for row in rows) for subject in SUBJECTS
            },
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source",
        type=Path,
        default=Path.home() / "Downloads" / "채점 결과",
        help="Directory containing grading result xlsx files.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().with_name("geo_cut_model.json"),
        help="Output JSON model path.",
    )
    parser.add_argument(
        "--ebsi",
        type=Path,
        default=Path(__file__).resolve().with_name("ebsi_geo_data.json"),
        help="Optional EBSi cache JSON from fetch_ebsi_geo_data.py.",
    )
    parser.add_argument(
        "--ebsi-months",
        default=",".join(DEFAULT_EBSI_MONTHS),
        help="Comma-separated EBSi months to train on, or 'all'. Default: 06,09,11.",
    )
    args = parser.parse_args()

    ebsi_months = None
    if args.ebsi_months.lower() != "all":
        ebsi_months = {
            month.strip().zfill(2)
            for month in args.ebsi_months.split(",")
            if month.strip()
        }

    model = build_model(args.source, args.ebsi, ebsi_months)
    args.output.write_text(json.dumps(model, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {args.output}")
    print(json.dumps({cut: model["cut_models"][cut]["cross_validation"] for cut in CUTS}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
